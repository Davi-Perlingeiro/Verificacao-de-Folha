"""
Exportador Excel para a folha de pagamento processada.
Gera planilha no formato do DP com formatacao profissional.
"""
import io
import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


# Headers da planilha de saida
HEADERS = [
    'Evento/Local',         # A
    'Data',                 # B
    'Nome',                 # C
    'Funcao Exercida',      # D
    'Horario Inicial',      # E
    'Horario Final',        # F
    'Funcao Contrato',      # G
    'Salario Hora',         # H
    'Sal. Funcao',          # I
    'Gratificacao',         # J
    'Total Horas',          # K
    'Intervalo',            # L
    'Horas Trabalhadas',    # M
    'Ad. Noturno (h)',      # N
    'Horas Trab. (num)',    # O
    'Ad. Not. (num)',       # P
    'Salario',              # Q
    'Gratif. Funcao',       # R
    'Ad. Not. 20%',         # S
    'DSR',                  # T
    'Ferias',               # U
    'Adic. Ferias',         # V
    '13o Salario',          # W
    'INSS',                 # X
    'INSS 13o',             # Y
    'Desc. VT',             # Z
    'Sal. Liquido',         # AA
    'Ajuda de Custo',       # AB
    'Desc. Adiant.',        # AC
    'Sal. Familia',         # AD
    'Total a Pagar',        # AE
]

# Mapeamento coluna DataFrame -> indice da planilha (1-based)
COL_MAP = {
    'evento': 1,
    'data': 2,
    'nome': 3,
    'funcao': 4,
    'horario_inicial': 5,
    'horario_final': 6,
    'funcao_contrato': 7,
    'salario_hora': 8,
    'salario_funcao': 9,
    'gratificacao': 10,
    'total_horas': 11,
    'intervalo': 12,
    'horas_trabalhadas': 13,
    'horas_noturnas': 14,
    'horas_trab_num': 15,
    'horas_not_num': 16,
    'salario': 17,
    'gratif_funcao': 18,
    'ad_noturno_valor': 19,
    'dsr': 20,
    'ferias': 21,
    'adic_ferias': 22,
    'decimo': 23,
    'inss': 24,
    'inss_13': 25,
    'desc_vt': 26,
    'sal_liquido': 27,
    'ajuda_custo': 28,
    'desc_adiantamento': 29,
    'sal_familia': 30,
    'total_pagar': 31,
}

# Colunas monetarias (formato R$)
MONEY_COLS = {8, 9, 10, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31}

# Colunas de horas (formato numerico)
HOUR_COLS = {11, 12, 13, 14, 15, 16}


def export_payroll_excel(df: pd.DataFrame, params: dict, periodo: str = None) -> bytes:
    """
    Exporta DataFrame de folha para Excel formatado.

    Args:
        df: DataFrame com os calculos de folha
        params: Parametros usados nos calculos
        periodo: Nome do periodo (ex: "09.02 A 15.02")

    Returns:
        bytes do arquivo Excel
    """
    wb = Workbook()
    ws = wb.active

    # Nome da sheet
    if periodo:
        ws.title = periodo
    else:
        ws.title = _detect_periodo(df)

    # Estilos
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=10, color='FFFFFF')
    param_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    money_fmt = '#,##0.00'
    hour_fmt = '0.00'

    # === ROW 1: Totais (sera preenchida depois) ===
    # Placeholder - preenchemos apos os dados

    # === ROW 2: Parametros ===
    sal_hora = params.get('salario_base_mensal', 1621) / params.get('divisor_horas', 220)
    param_values = {
        1: f"Sal.Base: R${params.get('salario_base_mensal', 1621):.2f}",
        8: sal_hora,
        20: params.get('dsr_fator', 1/6),
        21: params.get('ferias_fator', 1/12),
        22: params.get('ferias_adicional', 1/3),
        23: params.get('decimo_fator', 1/12),
        24: 'INSS Progressivo',
        25: 'INSS Progressivo',
        26: params.get('vt_desconto', 0),
    }
    for col, val in param_values.items():
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = param_fill
        cell.font = Font(italic=True, size=9)
        if isinstance(val, float) and val < 1:
            cell.number_format = '0.00%'
        elif isinstance(val, float):
            cell.number_format = money_fmt

    # === ROW 3: Headers ===
    for i, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # === ROWS 4+: Dados ===
    data_start_row = 4
    for row_idx, (_, row) in enumerate(df.iterrows(), data_start_row):
        for field, col_idx in COL_MAP.items():
            val = row.get(field, '')

            # Formatar datas
            if field == 'data' and isinstance(val, (datetime.date, datetime.datetime)):
                val = val.strftime('%d/%m/%Y')
            # Formatar horarios
            elif field in ('horario_inicial', 'horario_final'):
                if isinstance(val, datetime.time):
                    val = val.strftime('%H:%M')
                elif val:
                    val = str(val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

            # Formatacao numerica
            if col_idx in MONEY_COLS and isinstance(val, (int, float)):
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal='right')
            elif col_idx in HOUR_COLS and isinstance(val, (int, float)):
                cell.number_format = hour_fmt
                cell.alignment = Alignment(horizontal='center')

    last_data_row = data_start_row + len(df) - 1

    # === ROW 1: Totais com SUBTOTAL ===
    total_label_cell = ws.cell(row=1, column=1, value='TOTAIS')
    total_label_cell.font = Font(bold=True, size=11)
    total_label_cell.fill = total_fill

    # Colunas para totalizar
    total_cols = list(MONEY_COLS) + [15, 16]  # monetarias + horas numericas
    for col_idx in total_cols:
        col_letter = get_column_letter(col_idx)
        formula = f'=SUBTOTAL(9,{col_letter}{data_start_row}:{col_letter}{last_data_row})'
        cell = ws.cell(row=1, column=col_idx, value=formula)
        cell.font = Font(bold=True, size=10)
        cell.fill = total_fill
        cell.border = thin_border
        if col_idx in MONEY_COLS:
            cell.number_format = money_fmt
        else:
            cell.number_format = hour_fmt

    # === Ajustar largura das colunas ===
    col_widths = {
        1: 25, 2: 12, 3: 30, 4: 30, 5: 10, 6: 10, 7: 25,
        8: 10, 9: 10, 10: 10, 11: 8, 12: 8, 13: 10, 14: 10,
        15: 10, 16: 10, 17: 12, 18: 12, 19: 12, 20: 12,
        21: 12, 22: 12, 23: 12, 24: 12, 25: 12, 26: 10,
        27: 14, 28: 12, 29: 12, 30: 12, 31: 14,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Congelar paineis (headers visiveis ao rolar)
    ws.freeze_panes = 'A4'

    # Salvar em bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _detect_periodo(df: pd.DataFrame) -> str:
    """Detecta o periodo dos dados para nomear a sheet."""
    if df.empty or 'data' not in df.columns:
        return 'Folha'

    dates = pd.to_datetime(df['data'], errors='coerce').dropna()
    if dates.empty:
        return 'Folha'

    min_date = dates.min()
    max_date = dates.max()
    return f"{min_date.strftime('%d.%m')} A {max_date.strftime('%d.%m')}"
